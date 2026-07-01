from __future__ import annotations

import hashlib
import io
import csv
import uuid
from collections import Counter
from typing import Annotated, Any
from urllib.parse import unquote

from openpyxl import load_workbook
from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from api import models, schemas
from api.database import get_db
from api.services import (
    OM_COLUMN_MAP,
    apply_patch_to_asset,
    asset_code_index,
    audit_for_entity,
    model_summary,
    om_values,
    run_validation,
    serialize_asset,
    write_audit,
)
router = APIRouter(prefix="/api/v1")
Db = Annotated[Session, Depends(get_db)]

BMS_COLUMNS = [
    "AssetCode",
    "BMSDeviceID",
    "DeviceName",
    "Status",
    "Floor",
    "Room",
    "Location",
    "Link",
    "Document",
    "Manufacturer",
]
BMS_COLUMN_ALIASES = {
    "AssetCode": {"assetcode", "asset_code", "asset code"},
    "BMSDeviceID": {"bmsdeviceid", "bms_device_id", "bms device id", "deviceid", "device_id"},
    "DeviceName": {"devicename", "device_name", "device name", "name"},
    "Status": {"status", "device_status", "device status"},
    "Floor": {"floor", "level", "storey"},
    "Room": {"room", "room_zone", "room zone", "zone"},
    "Location": {"location", "vsf.location"},
    "Link": {"link", "vsf.link", "bms_link"},
    "Document": {"document", "vsf.document", "manual", "documentation"},
    "Manufacturer": {"manufacturer", "vendor", "make"},
}


def actor_name(x_actor_name: Annotated[str | None, Header()] = None) -> str:
    actor = unquote(str(x_actor_name or "")).strip()
    if not actor:
        raise HTTPException(400, "Thiếu X-Actor-Name. Hãy nhập tên người thao tác.")
    return actor[:255]


Actor = Annotated[str, Depends(actor_name)]


def request_id(request: Request) -> str:
    return request.headers.get("X-Request-ID") or str(uuid.uuid4())


def get_model(db: Session, model_id: uuid.UUID) -> models.TwinModel:
    model = db.get(models.TwinModel, model_id)
    if not model:
        raise HTTPException(404, "Không tìm thấy model.")
    return model


def get_asset(db: Session, asset_id: uuid.UUID) -> models.Asset:
    asset = db.scalar(
        select(models.Asset)
        .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
        .where(models.Asset.id == asset_id)
    )
    if not asset:
        raise HTTPException(404, "Không tìm thấy asset.")
    return asset


@router.get("/health")
def health(db: Db) -> dict[str, str]:
    db.scalar(select(1))
    return {"status": "ok", "database": "connected"}


@router.get("/models")
def list_models(
    db: Db,
    limit: int = Query(100, ge=1, le=500),
) -> dict[str, Any]:
    saved_models = db.scalars(
        select(models.TwinModel)
        .options(selectinload(models.TwinModel.project))
        .where(models.TwinModel.active.is_(True))
        .order_by(models.TwinModel.imported_at.desc())
        .limit(limit)
    ).all()
    return {
        "items": [
            {
                "id": str(model.id),
                "projectId": str(model.project_id),
                "projectCode": model.project.code,
                "projectName": model.project.name,
                "sourceFile": model.source_file,
                "apsUrn": model.aps_urn,
                "validationProfile": model.validation_profile,
                "importedAt": model.imported_at.isoformat(),
                "active": model.active,
                "summary": model_summary(db, model.id),
            }
            for model in saved_models
        ]
    }


@router.post("/projects")
def create_project(payload: schemas.ProjectCreate, db: Db, actor: Actor, request: Request) -> dict[str, Any]:
    project = db.scalar(select(models.Project).where(models.Project.code == payload.code.strip()))
    if project:
        before = {"code": project.code, "name": project.name}
        project.name = payload.name.strip()
        if before["name"] != project.name:
            write_audit(
                db,
                model_id=None,
                entity_type="project",
                entity_id=str(project.id),
                action="project.updated",
                actor=actor,
                source="streamlit",
                request_id=request_id(request),
                before=before,
                after={"code": project.code, "name": project.name},
            )
    else:
        project = models.Project(code=payload.code.strip(), name=payload.name.strip())
        db.add(project)
        db.flush()
        write_audit(
            db,
            model_id=None,
            entity_type="project",
            entity_id=str(project.id),
            action="project.created",
            actor=actor,
            source="streamlit",
            request_id=request_id(request),
            after={"code": project.code, "name": project.name},
        )
    db.commit()
    return {"id": str(project.id), "code": project.code, "name": project.name}


@router.post("/projects/{project_id}/models")
def create_model(
    project_id: uuid.UUID,
    payload: schemas.ModelCreate,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    if not db.get(models.Project, project_id):
        raise HTTPException(404, "Không tìm thấy project.")
    model = db.scalar(
        select(models.TwinModel).where(
            models.TwinModel.project_id == project_id,
            models.TwinModel.source_file == payload.source_file,
            models.TwinModel.source_checksum == payload.source_checksum,
        )
    )
    if not model:
        model = models.TwinModel(project_id=project_id, **payload.model_dump())
        db.add(model)
        db.flush()
        write_audit(
            db,
            model_id=model.id,
            entity_type="model",
            entity_id=str(model.id),
            action="model.created",
            actor=actor,
            source="streamlit",
            request_id=request_id(request),
            after=payload.model_dump(),
        )
    else:
        before = {
            "aps_urn": model.aps_urn,
            "compliance_summary": model.compliance_summary,
        }
        model.aps_urn = payload.aps_urn or model.aps_urn
        model.compliance_summary = payload.compliance_summary
        after = {
            "aps_urn": model.aps_urn,
            "compliance_summary": model.compliance_summary,
        }
        if before != after:
            write_audit(
                db,
                model_id=model.id,
                entity_type="model",
                entity_id=str(model.id),
                action="model.updated",
                actor=actor,
                source="streamlit",
                request_id=request_id(request),
                before=before,
                after=after,
            )
    db.commit()
    return {"id": str(model.id), "projectId": str(model.project_id), "sourceFile": model.source_file}


@router.post("/models/{model_id}/objects:batch-upsert")
def upsert_objects(
    model_id: uuid.UUID,
    payload: schemas.ObjectBatch,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, int]:
    get_model(db, model_id)
    guids = [row.ifc_guid.strip() for row in payload.objects]
    if len(set(guids)) != len(guids):
        raise HTTPException(400, "Batch có IFC GlobalId bị trùng.")
    existing_assets = db.scalars(
        select(models.Asset)
        .join(models.IfcObject)
        .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
        .where(models.IfcObject.model_id == model_id, models.IfcObject.ifc_guid.in_(guids))
    ).all()
    assets_by_guid = {asset.ifc_object.ifc_guid: asset for asset in existing_assets}
    created = 0
    updated = 0
    for row in payload.objects:
        asset = assets_by_guid.get(row.ifc_guid.strip())
        if not asset:
            obj = models.IfcObject(model_id=model_id, ifc_guid=row.ifc_guid.strip())
            asset = models.Asset(ifc_object=obj)
            asset.om = models.AssetOm(model_id=model_id)
            db.add(asset)
            created += 1
        else:
            obj = asset.ifc_object
            updated += 1
        obj.name = row.name
        obj.ifc_class = row.ifc_class
        obj.object_type = row.object_type
        obj.floor = row.floor
        obj.room = row.room
        obj.location = row.location
        obj.raw_source = row.raw_source
        if asset.scope_source != "manual_approval":
            asset.operational_scope = row.operational_scope
            asset.scope_reason = row.scope_reason
            asset.scope_source = row.scope_source
        current_sources = dict(asset.om.field_sources or {})
        for field, column in OM_COLUMN_MAP.items():
            if current_sources.get(field) == "manual_approval":
                continue
            setattr(asset.om, column, str(row.om_values.get(field) or "").strip())
            current_sources[field] = row.field_sources.get(field, "missing")
        asset.om.field_sources = current_sources
    write_audit(
        db,
        model_id=model_id,
        entity_type="model",
        entity_id=str(model_id),
        action="ifc.batch_upserted",
        actor=actor,
        source="streamlit",
        request_id=request_id(request),
        after={"created": created, "updated": updated, "batchSize": len(payload.objects)},
    )
    db.commit()
    return {"created": created, "updated": updated, "processed": len(payload.objects)}


@router.get("/models/{model_id}/viewer-summary")
def viewer_summary(model_id: uuid.UUID, db: Db) -> dict[str, Any]:
    model = get_model(db, model_id)
    return {
        "model": {
            "id": str(model.id),
            "sourceFile": model.source_file,
            "apsUrn": model.aps_urn,
            "validationProfile": model.validation_profile,
        },
        "summary": model_summary(db, model_id),
    }


@router.get("/models/{model_id}/assets")
def list_assets(
    model_id: uuid.UUID,
    db: Db,
    scope: str = "",
    readiness: str = "",
    q: str = "",
    limit: int = Query(500, ge=1, le=50000),
    offset: int = Query(0, ge=0),
) -> dict[str, Any]:
    get_model(db, model_id)
    stmt = (
        select(models.Asset)
        .join(models.IfcObject)
        .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
        .where(models.IfcObject.model_id == model_id)
    )
    if scope:
        stmt = stmt.where(models.Asset.operational_scope == scope)
    if q.strip():
        pattern = f"%{q.strip()}%"
        stmt = stmt.outerjoin(models.AssetOm).where(
            or_(
                models.IfcObject.name.ilike(pattern),
                models.IfcObject.ifc_guid.ilike(pattern),
                models.AssetOm.emsd_asset_code.ilike(pattern),
                models.AssetOm.vsf_asset_code.ilike(pattern),
            )
        )
    all_rows = db.scalars(stmt.order_by(models.IfcObject.name).offset(offset).limit(limit + 1)).all()
    has_more = len(all_rows) > limit
    items = [serialize_asset(db, asset, include_bms=False) for asset in all_rows[:limit]]
    if readiness:
        items = [item for item in items if item["readinessStatus"].lower() == readiness.lower()]
    return {"items": items, "offset": offset, "limit": limit, "hasMore": has_more}


@router.get("/models/{model_id}/assets/by-ifc-guid/{ifc_guid}")
def asset_by_guid(model_id: uuid.UUID, ifc_guid: str, db: Db) -> dict[str, Any]:
    asset = db.scalar(
        select(models.Asset)
        .join(models.IfcObject)
        .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
        .where(models.IfcObject.model_id == model_id, func.lower(models.IfcObject.ifc_guid) == ifc_guid.lower())
    )
    if not asset:
        raise HTTPException(404, "IFC GlobalId chưa có trong database.")
    return serialize_asset(db, asset, include_audit=True)


@router.post("/assets/{asset_id}/changes")
def create_change(
    asset_id: uuid.UUID,
    payload: schemas.ChangeCreate,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    asset = get_asset(db, asset_id)
    if payload.base_version != asset.row_version:
        raise HTTPException(409, "Dữ liệu đã thay đổi. Hãy tải lại asset trước khi sửa.")
    change = models.ChangeRequest(
        model_id=asset.ifc_object.model_id,
        entity_id=asset.id,
        patch=payload.patch,
        base_version=payload.base_version,
        created_by=actor,
    )
    db.add(change)
    db.flush()
    write_audit(
        db,
        model_id=asset.ifc_object.model_id,
        entity_type="change_request",
        entity_id=str(change.id),
        action="change_request.created",
        actor=actor,
        source=payload.source,
        request_id=request_id(request),
        after={"assetId": str(asset.id), "patch": payload.patch, "baseVersion": payload.base_version},
    )
    db.commit()
    return {"id": str(change.id), "status": change.status, "asset": serialize_asset(db, asset)}


def check_asset_codes_unique(db: Session, asset: models.Asset, values: dict[str, str]) -> None:
    for field, column in (
        ("EMSD.Common.Asset Code", models.AssetOm.emsd_asset_code),
        ("VSF.Common.Asset Code", models.AssetOm.vsf_asset_code),
    ):
        code = str(values.get(field) or "").strip()
        if not code:
            continue
        duplicate = db.scalar(
            select(models.AssetOm.asset_id).where(
                models.AssetOm.model_id == asset.ifc_object.model_id,
                func.upper(column) == code.upper(),
                models.AssetOm.asset_id != asset.id,
            )
        )
        if duplicate:
            raise HTTPException(409, f"{field} '{code}' đang thuộc một asset khác trong model.")


@router.post("/change-requests/{change_id}/approve")
def approve_change(
    change_id: uuid.UUID,
    payload: schemas.DecisionRequest,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    change = db.get(models.ChangeRequest, change_id)
    if not change:
        raise HTTPException(404, "Không tìm thấy bản nháp.")
    if change.status != "draft":
        raise HTTPException(409, "Bản nháp này đã được xử lý.")
    asset = get_asset(db, change.entity_id)
    if asset.row_version != change.base_version:
        raise HTTPException(409, "Asset đã thay đổi sau khi tạo bản nháp. Hãy tạo lại bản sửa.")
    proposed = {**om_values(asset.om), **(change.patch.get("values") or {})}
    check_asset_codes_unique(db, asset, proposed)
    try:
        before, after = apply_patch_to_asset(asset, change.patch)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    change.status = "approved"
    change.decided_by = actor
    change.decided_at = models.utcnow()
    write_audit(
        db,
        model_id=change.model_id,
        entity_type="asset",
        entity_id=str(asset.id),
        action="asset.change_approved",
        actor=actor,
        source=payload.source,
        request_id=request_id(request),
        before=before,
        after=after,
    )
    run_validation(db, change.model_id, actor, "vsf_om_10")
    db.commit()
    return {"changeRequestId": str(change.id), "status": change.status, "asset": serialize_asset(db, asset)}


@router.post("/change-requests/{change_id}/reject")
def reject_change(
    change_id: uuid.UUID,
    payload: schemas.DecisionRequest,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    change = db.get(models.ChangeRequest, change_id)
    if not change:
        raise HTTPException(404, "Không tìm thấy bản nháp.")
    if change.status != "draft":
        raise HTTPException(409, "Bản nháp này đã được xử lý.")
    change.status = "rejected"
    change.decided_by = actor
    change.decided_at = models.utcnow()
    write_audit(
        db,
        model_id=change.model_id,
        entity_type="change_request",
        entity_id=str(change.id),
        action="change_request.rejected",
        actor=actor,
        source=payload.source,
        request_id=request_id(request),
        before={"status": "draft"},
        after={"status": "rejected", "reason": payload.reason},
    )
    db.commit()
    return {"id": str(change.id), "status": change.status}


@router.post("/models/{model_id}/validation-runs")
def create_validation_run(
    model_id: uuid.UUID,
    payload: schemas.ValidationRunCreate,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    get_model(db, model_id)
    run = run_validation(db, model_id, actor, payload.profile)
    write_audit(
        db,
        model_id=model_id,
        entity_type="validation_run",
        entity_id=str(run.id),
        action="validation.completed",
        actor=actor,
        source="streamlit",
        request_id=request_id(request),
        after=run.summary,
    )
    db.commit()
    return {"id": str(run.id), "summary": run.summary}


def read_bms_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        raw_rows = list(csv.DictReader(io.StringIO(text)))
    elif lower.endswith((".xlsx", ".xls")):
        if lower.endswith(".xls"):
            raise HTTPException(400, "Định dạng .xls cũ không được hỗ trợ; hãy lưu thành .xlsx hoặc CSV.")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raw_rows = []
        else:
            headers = [str(value or "").strip() for value in values[0]]
            raw_rows = [
                dict(zip(headers, row, strict=False))
                for row in values[1:]
                if any(value not in (None, "") for value in row)
            ]
    else:
        raise HTTPException(400, "File BMS phải là CSV hoặc Excel.")
    normalized_headers = {
        str(header or "").strip().lower(): header
        for row in raw_rows[:1]
        for header in row
    }
    column_map = {
        canonical: next(
            (
                normalized_headers[alias]
                for alias in aliases
                if alias in normalized_headers
            ),
            None,
        )
        for canonical, aliases in BMS_COLUMN_ALIASES.items()
    }
    rows = [
        {
            canonical: str(row.get(source) or "").strip() if source else ""
            for canonical, source in column_map.items()
        }
        for row in raw_rows
    ]
    if not any(row["AssetCode"] for row in rows):
        raise HTTPException(400, "File BMS phải có cột AssetCode và ít nhất một mã asset.")
    return rows


@router.post("/models/{model_id}/bms-imports")
async def import_bms(
    model_id: uuid.UUID,
    db: Db,
    actor: Actor,
    request: Request,
    file: UploadFile = File(...),
) -> dict[str, Any]:
    get_model(db, model_id)
    content = await file.read()
    rows = read_bms_upload(file.filename or "bms.csv", content)
    checksum = hashlib.sha256(content).hexdigest()
    existing_batch = db.scalar(
        select(models.BmsImportBatch).where(
            models.BmsImportBatch.model_id == model_id,
            models.BmsImportBatch.checksum == checksum,
        )
    )
    if existing_batch:
        existing_rows = db.scalars(
            select(models.BmsImportRow).where(models.BmsImportRow.batch_id == existing_batch.id)
        ).all()
        code_index = asset_code_index(db, model_id)
        code_counts = Counter(row.asset_code.strip().upper() for row in existing_rows)
        counts = Counter()
        for row in existing_rows:
            if row.reconciliation_status in {"applied", "rejected"}:
                counts[row.reconciliation_status] += 1
                continue
            code = row.asset_code.strip().upper()
            candidates = code_index.get(code, [])
            if not row.bms_device_id:
                status, error = "pending_invalid", "Thiếu BMS Device ID."
            elif code_counts[code] > 1:
                status, error = "pending_duplicate_bms", "AssetCode xuất hiện nhiều dòng trong file BMS."
            elif not candidates:
                status, error = "pending_unmatched", "Không tìm thấy AssetCode tương ứng trong IFC."
            elif len(candidates) > 1:
                status, error = "pending_duplicate_ifc", "AssetCode thuộc nhiều object IFC."
            else:
                status, error = "auto_ready", ""
            row.reconciliation_status = status
            row.candidate_asset_ids = [str(value) for value in candidates]
            row.error = error
            counts[status] += 1
        existing_batch.counts = dict(counts)
        write_audit(
            db,
            model_id=model_id,
            entity_type="bms_import_batch",
            entity_id=str(existing_batch.id),
            action="bms.reconciled",
            actor=actor,
            source="streamlit",
            request_id=request_id(request),
            after={"counts": dict(counts)},
        )
        db.commit()
        return {
            "id": str(existing_batch.id),
            "filename": existing_batch.filename,
            "rows": len(existing_rows),
            "counts": existing_batch.counts,
            "existing": True,
        }
    code_index = asset_code_index(db, model_id)
    code_counts = Counter(str(row.get("AssetCode") or "").strip().upper() for row in rows)
    batch = models.BmsImportBatch(
        model_id=model_id,
        filename=file.filename or "bms.csv",
        checksum=checksum,
        uploaded_by=actor,
    )
    db.add(batch)
    db.flush()
    counts = Counter()
    for index, row in enumerate(rows):
        clean_row = {key: str(value or "").strip() for key, value in row.items()}
        code = clean_row["AssetCode"].upper()
        candidates = code_index.get(code, [])
        if not clean_row["BMSDeviceID"]:
            status = "pending_invalid"
            error = "Thiếu BMS Device ID."
        elif code_counts[code] > 1:
            status = "pending_duplicate_bms"
            error = "AssetCode xuất hiện nhiều dòng trong file BMS."
        elif not candidates:
            status = "pending_unmatched"
            error = "Không tìm thấy AssetCode tương ứng trong IFC."
        elif len(candidates) > 1:
            status = "pending_duplicate_ifc"
            error = "AssetCode thuộc nhiều object IFC."
        else:
            status = "auto_ready"
            error = ""
        counts[status] += 1
        db.add(
            models.BmsImportRow(
                batch_id=batch.id,
                row_number=index + 2,
                asset_code=clean_row["AssetCode"],
                bms_device_id=clean_row["BMSDeviceID"],
                payload=clean_row,
                reconciliation_status=status,
                candidate_asset_ids=[str(value) for value in candidates],
                error=error,
            )
        )
    batch.counts = dict(counts)
    write_audit(
        db,
        model_id=model_id,
        entity_type="bms_import_batch",
        entity_id=str(batch.id),
        action="bms.imported",
        actor=actor,
        source="streamlit",
        request_id=request_id(request),
        after={"filename": batch.filename, "rows": len(rows), "counts": dict(counts)},
    )
    db.commit()
    return {"id": str(batch.id), "filename": batch.filename, "rows": len(rows), "counts": dict(counts)}


def bms_row_payload(
    row: models.BmsImportRow,
    candidate_assets: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    candidate_assets = candidate_assets or {}
    return {
        "id": str(row.id),
        "row": row.row_number,
        "assetCode": row.asset_code,
        "bmsDeviceId": row.bms_device_id,
        "status": row.reconciliation_status,
        "candidateAssetIds": row.candidate_asset_ids,
        "candidateAssets": [
            candidate_assets[asset_id]
            for asset_id in row.candidate_asset_ids
            if asset_id in candidate_assets
        ],
        "error": row.error,
        "payload": row.payload,
    }


@router.get("/bms-imports/{batch_id}/reconciliation")
def bms_reconciliation(batch_id: uuid.UUID, db: Db) -> dict[str, Any]:
    batch = db.get(models.BmsImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "Không tìm thấy lần import BMS.")
    rows = db.scalars(
        select(models.BmsImportRow).where(models.BmsImportRow.batch_id == batch_id).order_by(models.BmsImportRow.row_number)
    ).all()
    candidate_ids = {
        uuid.UUID(asset_id)
        for row in rows
        for asset_id in row.candidate_asset_ids
    }
    candidates = (
        db.scalars(
            select(models.Asset)
            .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
            .where(models.Asset.id.in_(candidate_ids))
        ).all()
        if candidate_ids
        else []
    )
    candidate_assets = {
        str(asset.id): {
            "id": str(asset.id),
            "ifcGuid": asset.ifc_object.ifc_guid,
            "name": asset.ifc_object.name,
            "ifcClass": asset.ifc_object.ifc_class,
            "floor": asset.ifc_object.floor,
            "room": asset.ifc_object.room,
            "operationalScope": asset.operational_scope,
        }
        for asset in candidates
    }
    return {
        "batchId": str(batch.id),
        "counts": batch.counts,
        "items": [bms_row_payload(row, candidate_assets) for row in rows],
    }


def apply_bms_row(
    db: Session,
    row: models.BmsImportRow,
    asset: models.Asset,
    actor: str,
    decision: str,
    reason: str,
    source: str,
    request: Request,
) -> None:
    existing = db.scalar(
        select(models.BmsDevice).where(
            models.BmsDevice.model_id == asset.ifc_object.model_id,
            or_(models.BmsDevice.asset_id == asset.id, models.BmsDevice.bms_device_id == row.bms_device_id),
        )
    )
    if existing and (existing.asset_id != asset.id or existing.bms_device_id != row.bms_device_id):
        raise HTTPException(409, "BMS Device ID hoặc asset đã được map.")
    payload = row.payload
    location = payload.get("Location") or " / ".join(
        value for value in (payload.get("Floor"), payload.get("Room")) if value
    )
    before = {"operationalScope": asset.operational_scope, **om_values(asset.om)}
    if not existing:
        existing = models.BmsDevice(
            model_id=asset.ifc_object.model_id,
            asset_id=asset.id,
            bms_device_id=row.bms_device_id,
            source_batch_id=row.batch_id,
        )
        db.add(existing)
    existing.device_name = payload.get("DeviceName", "")
    existing.status = payload.get("Status", "")
    existing.floor = payload.get("Floor", "")
    existing.room = payload.get("Room", "")
    existing.location = location
    existing.link = payload.get("Link", "")
    existing.document = payload.get("Document", "")
    existing.manufacturer = payload.get("Manufacturer", "")
    bms_values = {
        "EMSD.Common.Asset Code": row.asset_code,
        "VSF.Common.Asset Code": row.asset_code,
        "VSF.Status": payload.get("Status", ""),
        "VSF.Location": location,
        "VSF.Link": payload.get("Link", ""),
        "VSF.Document": payload.get("Document", ""),
        "EMSD.Common.Manufacturer": payload.get("Manufacturer", ""),
        "VSF.Common.Manufacturer": payload.get("Manufacturer", ""),
    }
    proposed_values = om_values(asset.om)
    for field in ("EMSD.Common.Asset Code", "VSF.Common.Asset Code"):
        if not proposed_values[field].strip():
            proposed_values[field] = row.asset_code
    check_asset_codes_unique(db, asset, proposed_values)
    sources = dict(asset.om.field_sources or {})
    for field, value in bms_values.items():
        current = str(getattr(asset.om, OM_COLUMN_MAP[field]) or "").strip()
        is_asset_code = field in {"EMSD.Common.Asset Code", "VSF.Common.Asset Code"}
        if value and sources.get(field) != "manual_approval" and (not is_asset_code or not current):
            setattr(asset.om, OM_COLUMN_MAP[field], value)
            sources[field] = "bms_device_register"
    asset.om.field_sources = sources
    asset.operational_scope = "realtime"
    asset.scope_source = "bms_device_register"
    asset.scope_reason = "AssetCode khớp với BMS Device Register"
    asset.row_version += 1
    row.reconciliation_status = "applied"
    db.add(
        models.BmsMappingDecision(
            import_row_id=row.id,
            asset_id=asset.id,
            decision=decision,
            decided_by=actor,
            reason=reason,
        )
    )
    write_audit(
        db,
        model_id=asset.ifc_object.model_id,
        entity_type="asset",
        entity_id=str(asset.id),
        action=f"bms.mapping_{decision}",
        actor=actor,
        source=source,
        request_id=request_id(request),
        before=before,
        after={"operationalScope": asset.operational_scope, **om_values(asset.om), "bmsDeviceId": row.bms_device_id},
    )


@router.post("/bms-imports/{batch_id}/auto-apply")
def auto_apply_bms(batch_id: uuid.UUID, db: Db, actor: Actor, request: Request) -> dict[str, int]:
    batch = db.get(models.BmsImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "Không tìm thấy lần import BMS.")
    rows = db.scalars(
        select(models.BmsImportRow).where(
            models.BmsImportRow.batch_id == batch_id,
            models.BmsImportRow.reconciliation_status == "auto_ready",
        )
    ).all()
    applied = 0
    for row in rows:
        if len(row.candidate_asset_ids) != 1:
            continue
        asset = get_asset(db, uuid.UUID(row.candidate_asset_ids[0]))
        apply_bms_row(db, row, asset, actor, "auto", "AssetCode khớp duy nhất", "bms_import", request)
        applied += 1
    run_validation(db, batch.model_id, actor, "vsf_om_10")
    db.commit()
    return {"requested": len(rows), "applied": applied}


@router.post("/bms-import-rows/{row_id}/confirm")
def confirm_bms(
    row_id: uuid.UUID,
    payload: schemas.BmsConfirmRequest,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    row = db.get(models.BmsImportRow, row_id)
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng BMS.")
    if row.reconciliation_status in {"applied", "rejected"}:
        raise HTTPException(409, "Dòng BMS đã được xử lý.")
    asset = get_asset(db, payload.target_asset_id)
    batch = db.get(models.BmsImportBatch, row.batch_id)
    if asset.ifc_object.model_id != batch.model_id:
        raise HTTPException(400, "Asset không thuộc model của file BMS.")
    apply_bms_row(db, row, asset, actor, "confirmed", payload.reason, payload.source, request)
    run_validation(db, batch.model_id, actor, "vsf_om_10")
    db.commit()
    return {"row": bms_row_payload(row), "asset": serialize_asset(db, asset)}


@router.post("/bms-import-rows/{row_id}/reject")
def reject_bms(
    row_id: uuid.UUID,
    payload: schemas.DecisionRequest,
    db: Db,
    actor: Actor,
    request: Request,
) -> dict[str, Any]:
    row = db.get(models.BmsImportRow, row_id)
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng BMS.")
    batch = db.get(models.BmsImportBatch, row.batch_id)
    previous_status = row.reconciliation_status
    row.reconciliation_status = "rejected"
    db.add(
        models.BmsMappingDecision(
            import_row_id=row.id,
            asset_id=None,
            decision="rejected",
            decided_by=actor,
            reason=payload.reason,
        )
    )
    write_audit(
        db,
        model_id=batch.model_id,
        entity_type="bms_import_row",
        entity_id=str(row.id),
        action="bms.mapping_rejected",
        actor=actor,
        source=payload.source,
        request_id=request_id(request),
        before={"status": previous_status},
        after={"status": "rejected", "reason": payload.reason},
    )
    db.commit()
    return bms_row_payload(row)


@router.get("/assets/{asset_id}/audit")
def asset_audit(asset_id: uuid.UUID, db: Db, limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
    get_asset(db, asset_id)
    return {"items": audit_for_entity(db, "asset", str(asset_id), limit)}


@router.get("/models/{model_id}/audit")
def model_audit(model_id: uuid.UUID, db: Db, limit: int = Query(200, ge=1, le=1000)) -> dict[str, Any]:
    get_model(db, model_id)
    events = db.scalars(
        select(models.AuditEvent)
        .where(models.AuditEvent.model_id == model_id)
        .order_by(models.AuditEvent.occurred_at.desc())
        .limit(limit)
    ).all()
    return {
        "items": [
            {
                "id": event.id,
                "entityType": event.entity_type,
                "entityId": event.entity_id,
                "action": event.action,
                "actor": event.actor_display_name,
                "source": event.source,
                "changes": event.changes,
                "occurredAt": event.occurred_at.isoformat(),
            }
            for event in events
        ]
    }


@router.get("/models/{model_id}/snapshot-export")
def snapshot_export(model_id: uuid.UUID, db: Db) -> JSONResponse:
    model = get_model(db, model_id)
    assets = db.scalars(
        select(models.Asset)
        .join(models.IfcObject)
        .options(selectinload(models.Asset.ifc_object), selectinload(models.Asset.om))
        .where(models.IfcObject.model_id == model_id)
    ).all()
    payload = {
        "schemaVersion": "2.0.0",
        "kind": "validated-digital-twin-snapshot",
        "readOnlyExport": True,
        "model": {"id": str(model.id), "sourceFile": model.source_file},
        "summary": model_summary(db, model_id),
        "assets": [serialize_asset(db, asset) for asset in assets],
    }
    return JSONResponse(
        payload,
        headers={"Content-Disposition": f'attachment; filename="{model.id}_validated_twin_export.json"'},
    )
